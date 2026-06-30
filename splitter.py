#!/usr/bin/env python3
"""
All Inlinks Splitter — splits large Screaming Frog Excel crawl exports
into manageable files by geographical region, URL pattern, or both.

Two-pass streaming architecture for 900+ MB files:
  Pass 1: read_only — count destination frequencies (no row storage)
  Pass 2: read_only → write_only — route rows to output workbooks
"""

import re
import sys
import argparse
from datetime import date
from pathlib import Path
from collections import Counter, defaultdict
from urllib.parse import urlparse

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

# ============================================================
# CONFIGURATION
# ============================================================

IGNORED_TYPES = ['Sitemap Hreflang', 'XML Sitemap']

HIGH_THRESHOLD = 100
MEDIUM_THRESHOLD = 10

REGION_ALIASES = {
    'APAC': ['APAC'],
    'MEISA': ['MEISA'],
    'EU': ['EU'],
    'LAC': ['LAC'],
    'US': ['US', 'USA', 'United States', 'United States of America'],
    'Canada': ['Canada'],
}
SEGMENT_TOKEN_RE = re.compile(r'[A-Za-z0-9]+')

# Locale-like path prefixes to skip when extracting URL groups
LOCALE_RE = re.compile(r'^[a-z]{2}(?:-[a-z]{2})?$', re.IGNORECASE)

# ============================================================
# PATH DETECTION
# ============================================================

SCRIPT_DIR = Path(__file__).parent.resolve()

def _detect_dirs(cli_input=None, cli_output=None):
    if cli_input:
        input_dir = Path(cli_input)
    elif Path('/app/input').exists():
        input_dir = Path('/app/input')
    else:
        input_dir = SCRIPT_DIR / 'input'

    if cli_output:
        output_base = Path(cli_output)
    elif Path('/app/output').exists():
        output_base = Path('/app/output')
    else:
        output_base = SCRIPT_DIR / 'output'

    return input_dir, output_base

# ============================================================
# PROGRESS TRACKER
# ============================================================

class ProgressTracker:
    def __init__(self, total_steps):
        self.total = total_steps
        self.current = 0

    def step(self, description):
        self.current += 1
        print(f"\n[Step {self.current} of {self.total}] {description}")
        print("-" * 50)

# ============================================================
# UTILITY FUNCTIONS
# ============================================================

def find_column_index(headers, column_name):
    """Find column index by name (case-insensitive substring match)."""
    for idx, header in enumerate(headers):
        if header and column_name.lower() in str(header).lower():
            return idx
    return None


def is_ignored_type(type_value):
    """Check if a row's Type should be filtered out."""
    if type_value is None:
        return False
    type_str = str(type_value).lower()
    return any(ignored.lower() in type_str for ignored in IGNORED_TYPES)


def get_priority(frequency):
    """Assign priority label based on destination frequency."""
    if frequency >= HIGH_THRESHOLD:
        return 'HIGH'
    elif frequency >= MEDIUM_THRESHOLD:
        return 'MEDIUM'
    return 'LOW'


def get_matching_regions(value):
    """Return list of regions matching the Source Segments value."""
    if value is None or str(value).strip() == '':
        return ['OTHER']
    segment_tokens = [token.upper() for token in SEGMENT_TOKEN_RE.findall(str(value))]
    matches = [
        region
        for region, aliases in REGION_ALIASES.items()
        if any(_matches_segment_alias(segment_tokens, alias) for alias in aliases)
    ]
    return matches if matches else ['OTHER']


def _matches_segment_alias(segment_tokens, alias):
    """Return True if alias appears as complete token(s) in Source Segments."""
    alias_tokens = [token.upper() for token in SEGMENT_TOKEN_RE.findall(alias)]
    if not alias_tokens:
        return False
    if len(alias_tokens) == 1:
        return alias_tokens[0] in segment_tokens
    end = len(segment_tokens) - len(alias_tokens) + 1
    return any(segment_tokens[i:i + len(alias_tokens)] == alias_tokens for i in range(end))


def extract_url_group(url, depth=2, pattern=None):
    """
    Derive a grouping key from a URL.

    If *pattern* is given, use it as a regex and return the first match group
    (or the full match if there are no groups).

    Otherwise strip the locale-like first segment (e.g. "en-us") and join the
    next *depth* path segments with "_".
    """
    if pattern:
        m = re.search(pattern, str(url))
        if m:
            return m.group(1) if m.lastindex else m.group(0)
        return 'other'

    try:
        path = urlparse(str(url)).path.strip('/')
    except Exception:
        return 'other'

    if not path:
        return 'root'

    # Strip file extension from last segment (e.g. .html, .php)
    path = re.sub(r'\.[a-zA-Z]{2,5}$', '', path)

    segments = path.split('/')

    # Skip locale-like prefix
    if segments and LOCALE_RE.match(segments[0]):
        segments = segments[1:]

    if not segments:
        return 'root'

    key = '_'.join(segments[:depth])
    # Sanitise for use in filenames
    key = re.sub(r'[^\w\-]', '_', key)
    return key if key else 'other'


def _safe_cell(row, idx):
    """Safely get a cell value from a row tuple."""
    if idx is not None and idx < len(row):
        return row[idx]
    return None

# ============================================================
# PASS 1 — ANALYSIS (read_only, no row storage)
# ============================================================

def analyze_workbook(xlsx_path, args):
    """
    Stream through every sheet with read_only=True.
    Returns (headers, dest_counter, bucket_counter, total_rows, kept_rows).

    headers: tuple from the first row of the first sheet
    dest_counter: Counter of destination URLs
    bucket_counter: Counter of bucket keys (for progress reporting)
    """
    wb = load_workbook(xlsx_path, read_only=True)
    headers = None
    type_col = None
    dest_col = None
    seg_col = None
    source_col = None
    dest_counter = Counter()
    bucket_counter = Counter()
    total_rows = 0
    kept_rows = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row_num, row in enumerate(ws.iter_rows(values_only=True)):
            if headers is None and row_num == 0:
                # First header row we encounter — use it
                headers = tuple(row)
                type_col = find_column_index(headers, 'Type')
                dest_col = find_column_index(headers, 'Destination')
                seg_col = find_column_index(headers, 'Source Segments')
                source_col = find_column_index(headers, 'Source')
                continue
            elif row_num == 0:
                # Subsequent sheet header row — skip
                continue

            # Skip empty rows
            if all(c is None for c in row):
                continue

            total_rows += 1

            # Optional type filtering
            if not args.no_filter and type_col is not None:
                if is_ignored_type(_safe_cell(row, type_col)):
                    continue

            kept_rows += 1

            # Destination counting
            dest_val = _safe_cell(row, dest_col)
            if dest_val:
                dest_counter[dest_val] += 1

            # Bucket counting
            buckets = _resolve_buckets(
                row, seg_col, source_col, args.split, args.url_depth, args.url_pattern
            )
            for b in buckets:
                bucket_counter[b] += 1

    wb.close()
    return headers, dest_counter, bucket_counter, total_rows, kept_rows


def _resolve_buckets(row, seg_col, source_col, split_mode, url_depth, url_pattern):
    """Determine which output bucket(s) a row belongs to."""
    if split_mode == 'region':
        return get_matching_regions(_safe_cell(row, seg_col))

    if split_mode == 'url':
        return [extract_url_group(_safe_cell(row, source_col), url_depth, url_pattern)]

    # both
    regions = get_matching_regions(_safe_cell(row, seg_col))
    url_group = extract_url_group(_safe_cell(row, source_col), url_depth, url_pattern)
    return [f"{r}_{url_group}" for r in regions]

# ============================================================
# OUTPUT MANAGER — lazy write_only workbooks
# ============================================================

class OutputManager:
    """
    Manages one write_only workbook per bucket.  Workbooks are created lazily
    on first row append so we never create empty files.
    """

    def __init__(self, output_dir, base_name, headers, dest_counter, no_summary):
        self.output_dir = output_dir
        self.base_name = base_name
        self.headers = list(headers) + ['Priority']
        self.dest_counter = dest_counter
        self.no_summary = no_summary
        self._workbooks = {}    # bucket → Workbook
        self._data_sheets = {}  # bucket → worksheet
        self._dest_sets = defaultdict(set)  # bucket → set of destinations
        self._row_counts = Counter()

    def _init_bucket(self, bucket):
        wb = Workbook(write_only=True)
        # Summary placeholder — we'll fill it at close() time
        ws_data = wb.create_sheet("Data")
        # Write header row
        ws_data.append(self.headers)
        self._workbooks[bucket] = wb
        self._data_sheets[bucket] = ws_data

    def append(self, bucket, row, dest_value, priority):
        if bucket not in self._workbooks:
            self._init_bucket(bucket)
        self._data_sheets[bucket].append(list(row) + [priority])
        self._row_counts[bucket] += 1
        if dest_value:
            self._dest_sets[bucket].add(dest_value)

    def close_all(self):
        """Write summary sheets (unless --no-summary) and save all workbooks."""
        files_created = 0
        for bucket, wb in self._workbooks.items():
            if not self.no_summary:
                self._write_summary(wb, bucket)
            filename = f"{self.base_name}_{bucket}.xlsx"
            wb.save(self.output_dir / filename)
            count = self._row_counts[bucket]
            dests = len(self._dest_sets[bucket])
            print(f"    {bucket}: {count:,} rows, {dests} unique destinations")
            files_created += 1
        return files_created

    def _write_summary(self, wb, bucket):
        """Create a Summary sheet at position 0 with destinations ranked by frequency."""
        # write_only sheets can only append; we create Summary and move it first
        ws = wb.create_sheet("Summary", 0)

        # Header row
        ws.append(['Priority', 'Destination', 'Occurrences', 'Impact'])

        dests_in_bucket = self._dest_sets[bucket]
        ranked = sorted(
            ((d, self.dest_counter[d]) for d in dests_in_bucket if d in self.dest_counter),
            key=lambda x: x[1],
            reverse=True,
        )
        for dest, freq in ranked:
            priority = get_priority(freq)
            impact = f"Fix 1 URL → removes {freq} error{'s' if freq > 1 else ''}"
            ws.append([priority, dest, freq, impact])

# ============================================================
# PASS 2 — SPLITTING (read_only → write_only)
# ============================================================

def split_workbook(xlsx_path, args, headers, dest_counter, output_dir):
    """
    Second streaming pass.  Routes each row to the correct OutputManager bucket,
    appending priority on the fly.
    """
    base_name = xlsx_path.stem
    om = OutputManager(output_dir, base_name, headers, dest_counter, args.no_summary)

    type_col = find_column_index(headers, 'Type')
    dest_col = find_column_index(headers, 'Destination')
    seg_col = find_column_index(headers, 'Source Segments')
    source_col = find_column_index(headers, 'Source')

    wb = load_workbook(xlsx_path, read_only=True)
    first_header_seen = False

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row_num, row in enumerate(ws.iter_rows(values_only=True)):
            if row_num == 0:
                if not first_header_seen:
                    first_header_seen = True
                # Skip header rows in every sheet
                continue

            if all(c is None for c in row):
                continue

            # Optional type filtering
            if not args.no_filter and type_col is not None:
                if is_ignored_type(_safe_cell(row, type_col)):
                    continue

            dest_val = _safe_cell(row, dest_col)
            freq = dest_counter.get(dest_val, 0) if dest_val else 0
            priority = get_priority(freq)

            buckets = _resolve_buckets(
                row, seg_col, source_col, args.split, args.url_depth, args.url_pattern
            )
            for bucket in buckets:
                om.append(bucket, row, dest_val, priority)

    wb.close()

    print(f"\n  Output files for {base_name}:")
    files_created = om.close_all()
    return files_created

# ============================================================
# CLI
# ============================================================

def build_parser():
    parser = argparse.ArgumentParser(
        description='Split large Screaming Frog All Inlinks exports into manageable files.',
    )
    parser.add_argument(
        '--split',
        choices=['region', 'url', 'both'],
        default='region',
        help='Splitting strategy (default: region)',
    )
    parser.add_argument(
        '--url-depth',
        type=int,
        default=2,
        metavar='N',
        help='Number of URL path segments for grouping in url/both mode (default: 2)',
    )
    parser.add_argument(
        '--url-pattern',
        type=str,
        default=None,
        metavar='REGEX',
        help='Custom regex for URL grouping (overrides --url-depth)',
    )
    parser.add_argument(
        '--no-summary',
        action='store_true',
        help='Skip creating the Summary sheet in output files',
    )
    parser.add_argument(
        '--no-filter',
        action='store_true',
        help='Do not filter out ignored types (Sitemap Hreflang, XML Sitemap)',
    )
    parser.add_argument(
        '--input',
        type=str,
        default=None,
        metavar='DIR',
        help='Input directory (default: input/ or /app/input in Docker)',
    )
    parser.add_argument(
        '--output',
        type=str,
        default=None,
        metavar='DIR',
        help='Output base directory (default: output/ or /app/output in Docker)',
    )
    return parser

# ============================================================
# MAIN
# ============================================================

def main():
    parser = build_parser()
    args = parser.parse_args()

    input_dir, output_base = _detect_dirs(args.input, args.output)
    input_dir.mkdir(parents=True, exist_ok=True)

    today = date.today().isoformat()
    output_dir = output_base / today
    output_dir.mkdir(parents=True, exist_ok=True)

    xlsx_files = sorted(input_dir.glob('*.xlsx'))
    if not xlsx_files:
        print(f"No .xlsx files found in {input_dir}")
        print("Place your Excel files in the input/ directory and run again.")
        sys.exit(1)

    # Each file needs 2 passes + final save → 3 steps per file
    total_steps = len(xlsx_files) * 3
    tracker = ProgressTracker(total_steps)

    print("=" * 60)
    print("All Inlinks Splitter")
    print("=" * 60)
    print(f"  Split mode : {args.split}")
    if args.split in ('url', 'both'):
        if args.url_pattern:
            print(f"  URL pattern: {args.url_pattern}")
        else:
            print(f"  URL depth  : {args.url_depth}")
    print(f"  Filtering  : {'off' if args.no_filter else 'on (excluding ' + ', '.join(IGNORED_TYPES) + ')'}")
    print(f"  Files      : {len(xlsx_files)}")
    print(f"  Output     : output/{today}/")

    grand_total_files = 0

    for xlsx_path in xlsx_files:
        # --- Pass 1: Analyse ---
        tracker.step(f"Analysing {xlsx_path.name}")
        try:
            headers, dest_counter, bucket_counter, total_rows, kept_rows = analyze_workbook(
                xlsx_path, args
            )
        except Exception as e:
            print(f"  ERROR: Could not read {xlsx_path.name} — {e}", file=sys.stderr)
            # Skip the remaining 2 steps for this file
            tracker.current += 2
            continue

        if headers is None:
            print(f"  WARNING: No data found in {xlsx_path.name}, skipping.")
            tracker.current += 2
            continue

        print(f"  Total rows : {total_rows:,}")
        print(f"  After filter: {kept_rows:,}")
        print(f"  Unique dests: {len(dest_counter):,}")
        print(f"  Buckets     : {len(bucket_counter)}")

        # --- Priority summary ---
        tracker.step(f"Computing priorities for {xlsx_path.name}")
        high = sum(1 for f in dest_counter.values() if f >= HIGH_THRESHOLD)
        med = sum(1 for f in dest_counter.values() if MEDIUM_THRESHOLD <= f < HIGH_THRESHOLD)
        low = sum(1 for f in dest_counter.values() if f < MEDIUM_THRESHOLD)
        print(f"  HIGH  : {high:,} destinations")
        print(f"  MEDIUM: {med:,} destinations")
        print(f"  LOW   : {low:,} destinations")

        # --- Pass 2: Split & write ---
        tracker.step(f"Splitting {xlsx_path.name}")
        files_created = split_workbook(xlsx_path, args, headers, dest_counter, output_dir)
        grand_total_files += files_created

    print("\n" + "=" * 60)
    print(f"Done. Created {grand_total_files} output file(s) in output/{today}/")
    print("=" * 60)


if __name__ == '__main__':
    main()
