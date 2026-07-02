import unittest
from argparse import Namespace
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook, load_workbook

from splitter import analyze_workbook, split_workbook


class SplitWorkbookRegionOutputTests(unittest.TestCase):
    def test_usa_source_segment_writes_usa_output_file(self):
        with TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            xlsx_path = tmp_path / 'crawl.xlsx'
            output_dir = tmp_path / 'output'
            output_dir.mkdir()

            wb = Workbook()
            ws = wb.active
            ws.title = 'All Inlinks'
            ws.append(['Type', 'Source', 'Destination', 'Source Segments'])
            ws.append(['Hyperlink', 'https://example.com/en-us/page', 'https://example.com/broken', 'USA'])
            wb.save(xlsx_path)

            args = Namespace(
                split='region',
                url_depth=2,
                url_pattern=None,
                no_filter=False,
                no_summary=False,
            )
            headers, dest_counter, bucket_counter, total_rows, kept_rows = analyze_workbook(xlsx_path, args)

            self.assertEqual(bucket_counter['USA'], 1)
            self.assertEqual(total_rows, 1)
            self.assertEqual(kept_rows, 1)

            files_created = split_workbook(xlsx_path, args, headers, dest_counter, output_dir)

            self.assertEqual(files_created, 1)
            self.assertTrue((output_dir / 'crawl_USA.xlsx').exists())
            self.assertFalse((output_dir / 'crawl_US.xlsx').exists())

            result = load_workbook(output_dir / 'crawl_USA.xlsx', read_only=True)
            try:
                self.assertEqual(result.sheetnames, ['Summary', 'Data'])
                data_rows = list(result['Data'].iter_rows(values_only=True))
                self.assertEqual(data_rows[1][-1], 'LOW')
            finally:
                result.close()


if __name__ == '__main__':
    unittest.main()
